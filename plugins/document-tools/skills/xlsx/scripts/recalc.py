"""
Excel Formula Recalculation Script
Recalculates all formulas using Excel COM (Windows) or LibreOffice (Linux/macOS)
"""

import json
import os
import platform
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook


def _recalc_xlwings(abs_path):
    import xlwings as xw
    app = xw.App(visible=False)
    try:
        wb = app.books.open(abs_path)
        app.calculate()
        wb.save()
        wb.close()
    finally:
        app.quit()


def _recalc_libreoffice(abs_path, timeout):
    from office.soffice import get_soffice_env, _find_soffice

    MACRO_DIR_MACOS = "~/Library/Application Support/LibreOffice/4/user/basic/Standard"
    MACRO_DIR_LINUX = "~/.config/libreoffice/4/user/basic/Standard"
    MACRO_FILENAME = "Module1.xba"
    RECALCULATE_MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""

    if platform.system() == "Darwin":
        macro_dir = os.path.expanduser(MACRO_DIR_MACOS)
    else:
        macro_dir = os.path.expanduser(MACRO_DIR_LINUX)
    macro_file = os.path.join(macro_dir, MACRO_FILENAME)

    if not (os.path.exists(macro_file) and "RecalculateAndSave" in Path(macro_file).read_text()):
        if not os.path.exists(macro_dir):
            subprocess.run(
                [_find_soffice(), "--headless", "--terminate_after_init"],
                capture_output=True, timeout=10, env=get_soffice_env(),
            )
            os.makedirs(macro_dir, exist_ok=True)
        Path(macro_file).write_text(RECALCULATE_MACRO)

    cmd = [
        _find_soffice(), "--headless", "--norestore",
        "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
        abs_path,
    ]

    if platform.system() == "Linux":
        cmd = ["timeout", str(timeout)] + cmd
    elif platform.system() == "Darwin":
        try:
            subprocess.run(["gtimeout", "--version"], capture_output=True, timeout=1, check=False)
            cmd = ["gtimeout", str(timeout)] + cmd
        except (FileNotFoundError, subprocess.TimeoutExpired):
            pass

    result = subprocess.run(cmd, capture_output=True, text=True, env=get_soffice_env())
    if result.returncode != 0 and result.returncode != 124:
        raise RuntimeError(result.stderr or "LibreOffice recalc failed")


def _scan_errors(filename):
    wb = load_workbook(filename, data_only=True)

    excel_errors = ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]
    error_details = {err: [] for err in excel_errors}
    total_errors = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and isinstance(cell.value, str):
                    for err in excel_errors:
                        if err in cell.value:
                            error_details[err].append(f"{sheet_name}!{cell.coordinate}")
                            total_errors += 1
                            break
    wb.close()

    wb_formulas = load_workbook(filename, data_only=False)
    formula_count = 0
    for sheet_name in wb_formulas.sheetnames:
        ws = wb_formulas[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1
    wb_formulas.close()

    result = {
        "status": "success" if total_errors == 0 else "errors_found",
        "total_errors": total_errors,
        "total_formulas": formula_count,
        "error_summary": {},
    }
    for err_type, locations in error_details.items():
        if locations:
            result["error_summary"][err_type] = {
                "count": len(locations),
                "locations": locations[:20],
            }
    return result


def recalc(filename, timeout=30):
    if not Path(filename).exists():
        return {"error": f"File {filename} does not exist"}

    abs_path = str(Path(filename).absolute())

    try:
        if platform.system() == "Windows":
            _recalc_xlwings(abs_path)
        else:
            _recalc_libreoffice(abs_path, timeout)
    except Exception as e:
        return {"error": f"Recalc failed: {e}"}

    try:
        return _scan_errors(filename)
    except Exception as e:
        return {"error": str(e)}


def main():
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <excel_file> [timeout_seconds]")
        print("\nRecalculates all formulas using Excel (Windows) or LibreOffice (Linux/macOS)")
        print("\nReturns JSON with error details:")
        print("  - status: 'success' or 'errors_found'")
        print("  - total_errors: Total number of Excel errors found")
        print("  - total_formulas: Number of formulas in the file")
        print("  - error_summary: Breakdown by error type with locations")
        sys.exit(1)

    filename = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30

    result = recalc(filename, timeout)
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
